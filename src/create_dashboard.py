import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, PieChart, ScatterChart, Reference, LineChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint, Series as ChartSeries
from openpyxl.utils import get_column_letter
import re

DATA_PATH = "data/yes24_it_mobile_bestsellers.csv"
OUTPUT_PATH = "data/yes24_dashboard.xlsx"

df = pd.read_csv(DATA_PATH)

def parse_year_month(date_str):
    if not isinstance(date_str, str):
        return None, None
    match = re.search(r'(\d{4})년\s*(\d{2})월', date_str)
    if match:
        return int(match.group(1)), int(match.group(2))
    return None, None

parsed = df['Publish Date'].apply(parse_year_month)
df['Publish Year'] = [p[0] for p in parsed]
df['Publish Month'] = [p[1] for p in parsed]

df['Discount Rate Numeric'] = df['Discount Rate'].astype(str).str.replace('%', '', regex=False)
df['Discount Rate Numeric'] = pd.to_numeric(df['Discount Rate Numeric'], errors='coerce').fillna(0)

wb = Workbook()

HEADER_FILL = PatternFill('solid', fgColor='2D3436')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(name='Arial', bold=True, size=16, color='2D3436')
SUBTITLE_FONT = Font(name='Arial', bold=True, size=12, color='636E72')
KPI_VALUE_FONT = Font(name='Arial', bold=True, size=20, color='6C5CE7')
KPI_LABEL_FONT = Font(name='Arial', size=10, color='636E72')
DATA_FONT = Font(name='Arial', size=10)
MONEY_FORMAT = '#,##0'
THIN_BORDER = Border(
    left=Side(style='thin', color='DFE6E9'),
    right=Side(style='thin', color='DFE6E9'),
    top=Side(style='thin', color='DFE6E9'),
    bottom=Side(style='thin', color='DFE6E9')
)
ALT_FILL = PatternFill('solid', fgColor='F8F9FA')

# =============================================
# Sheet 1: Dashboard
# =============================================
ws = wb.active
ws.title = "Dashboard"
ws.sheet_properties.tabColor = "6C5CE7"

ws.merge_cells('B2:I2')
ws['B2'] = 'YES24 IT/모바일 베스트셀러 대시보드'
ws['B2'].font = TITLE_FONT
ws['B2'].alignment = Alignment(horizontal='left')

ws.merge_cells('B3:I3')
ws['B3'] = '수집 데이터 기반 탐색적 데이터 분석 (EDA)'
ws['B3'].font = SUBTITLE_FONT

# KPI Cards Row
kpi_data = [
    ('B', 'C', '총 도서 수', f'=COUNTA(\'Raw Data\'!A2:A{len(df)+1})'),
    ('D', 'E', '평균 판매가', f'=ROUND(AVERAGE(\'Raw Data\'!F2:F{len(df)+1}),0)'),
    ('F', 'G', '평균 평점', f'=ROUND(AVERAGE(\'Raw Data\'!K2:K{len(df)+1}),2)'),
    ('H', 'I', '평균 할인율', f'=ROUND(AVERAGE(\'Raw Data\'!H2:H{len(df)+1}),1)'),
]

for start_col, end_col, label, formula in kpi_data:
    row = 5
    ws.merge_cells(f'{start_col}{row}:{end_col}{row}')
    cell_label = ws[f'{start_col}{row}']
    cell_label.value = label
    cell_label.font = KPI_LABEL_FONT
    cell_label.alignment = Alignment(horizontal='center')

    ws.merge_cells(f'{start_col}{row+1}:{end_col}{row+1}')
    cell_val = ws[f'{start_col}{row+1}']
    cell_val.value = formula
    cell_val.font = KPI_VALUE_FONT
    cell_val.alignment = Alignment(horizontal='center')
    if '가' in label:
        cell_val.number_format = '#,##0"원"'
    elif '평점' in label:
        cell_val.number_format = '0.00'
    elif '할인율' in label:
        cell_val.number_format = '0.0"%"'

# Publisher Top 10 Bar Chart
ws.merge_cells('B9:E9')
ws['B9'] = '출판사별 베스트셀러 수 TOP 10'
ws['B9'].font = Font(name='Arial', bold=True, size=12, color='2D3436')

pub_counts = df['Publisher'].value_counts().head(10).reset_index()
pub_counts.columns = ['Publisher', 'Count']
ws['B10'] = '출판사'
ws['C10'] = '도서 수'
ws['B10'].font = HEADER_FONT
ws['C10'].font = HEADER_FONT
ws['B10'].fill = HEADER_FILL
ws['C10'].fill = HEADER_FILL
ws['B10'].alignment = Alignment(horizontal='center')
ws['C10'].alignment = Alignment(horizontal='center')

for i, (_, row_data) in enumerate(pub_counts.iterrows()):
    r = 11 + i
    ws[f'B{r}'] = row_data['Publisher']
    ws[f'C{r}'] = row_data['Count']
    ws[f'B{r}'].font = DATA_FONT
    ws[f'C{r}'].font = DATA_FONT
    ws[f'C{r}'].alignment = Alignment(horizontal='center')
    if i % 2 == 1:
        ws[f'B{r}'].fill = ALT_FILL
        ws[f'C{r}'].fill = ALT_FILL

chart_bar = BarChart()
chart_bar.type = "bar"
chart_bar.style = 10
chart_bar.title = "출판사별 베스트셀러 수"
chart_bar.y_axis.title = "출판사"
chart_bar.x_axis.title = "도서 수"
chart_bar.width = 22
chart_bar.height = 14
cats = Reference(ws, min_col=2, min_row=11, max_row=20)
vals = Reference(ws, min_col=3, min_row=10, max_row=20)
chart_bar.add_data(vals, titles_from_data=True)
chart_bar.set_categories(cats)
chart_bar.shape = 4
ws.add_chart(chart_bar, "E9")

# Price Distribution Bar Chart
ws.merge_cells('B23:E23')
ws['B23'] = '가격대별 도서 분포'
ws['B23'].font = Font(name='Arial', bold=True, size=12, color='2D3436')

bins = [0, 15000, 20000, 25000, 30000, 35000, 100000]
labels_price = ['~1.5만', '1.5~2만', '2~2.5만', '2.5~3만', '3~3.5만', '3.5만~']
df['Price Range'] = pd.cut(df['Sale Price'], bins=bins, labels=labels_price)
price_dist = df['Price Range'].value_counts().reindex(labels_price).fillna(0)

ws['B24'] = '가격대'
ws['C24'] = '도서 수'
ws['B24'].font = HEADER_FONT
ws['C24'].font = HEADER_FONT
ws['B24'].fill = HEADER_FILL
ws['C24'].fill = HEADER_FILL
ws['B24'].alignment = Alignment(horizontal='center')
ws['C24'].alignment = Alignment(horizontal='center')

for i, label in enumerate(labels_price):
    r = 25 + i
    ws[f'B{r}'] = label
    ws[f'C{r}'] = int(price_dist[label])
    ws[f'B{r}'].font = DATA_FONT
    ws[f'C{r}'].font = DATA_FONT
    ws[f'C{r}'].alignment = Alignment(horizontal='center')
    if i % 2 == 1:
        ws[f'B{r}'].fill = ALT_FILL
        ws[f'C{r}'].fill = ALT_FILL

chart_price = BarChart()
chart_price.type = "col"
chart_price.style = 10
chart_price.title = "가격대별 도서 분포"
chart_price.y_axis.title = "도서 수"
chart_price.width = 22
chart_price.height = 14
cats_p = Reference(ws, min_col=2, min_row=25, max_row=25+len(labels_price)-1)
vals_p = Reference(ws, min_col=3, min_row=24, max_row=24+len(labels_price))
chart_price.add_data(vals_p, titles_from_data=True)
chart_price.set_categories(cats_p)
chart_price.shape = 4
ws.add_chart(chart_price, "E23")

# Rating vs Sale Index Scatter
ws.merge_cells('B40:E40')
ws['B40'] = '평점 vs 판매지수'
ws['B40'].font = Font(name='Arial', bold=True, size=12, color='2D3436')

# Scatter chart - will be added after Raw Data sheet is created
scatter_placeholder = True

# Publish Trend Line Chart
ws.merge_cells('B57:E57')
ws['B57'] = '출판 월별 트렌드'
ws['B57'].font = Font(name='Arial', bold=True, size=12, color='2D3436')

trend_data = df.groupby(['Publish Year', 'Publish Month']).size().reset_index(name='Count')
trend_data = trend_data.dropna(subset=['Publish Year', 'Publish Month'])
trend_data = trend_data.sort_values(['Publish Year', 'Publish Month'])
trend_data['Label'] = trend_data.apply(lambda r: f"{int(r['Publish Year'])}/{int(r['Publish Month']):02d}", axis=1)

ws['B58'] = '연월'
ws['C58'] = '출판 수'
ws['B58'].font = HEADER_FONT
ws['C58'].font = HEADER_FONT
ws['B58'].fill = HEADER_FILL
ws['C58'].fill = HEADER_FILL
ws['B58'].alignment = Alignment(horizontal='center')
ws['C58'].alignment = Alignment(horizontal='center')

for i, (_, row_data) in enumerate(trend_data.iterrows()):
    r = 59 + i
    ws[f'B{r}'] = row_data['Label']
    ws[f'C{r}'] = int(row_data['Count'])
    ws[f'B{r}'].font = DATA_FONT
    ws[f'C{r}'].font = DATA_FONT
    ws[f'C{r}'].alignment = Alignment(horizontal='center')

if len(trend_data) > 0:
    chart_line = LineChart()
    chart_line.title = "출판 월별 트렌드"
    chart_line.y_axis.title = "출판 수"
    chart_line.style = 10
    chart_line.width = 22
    chart_line.height = 14
    cats_l = Reference(ws, min_col=2, min_row=59, max_row=58+len(trend_data))
    vals_l = Reference(ws, min_col=3, min_row=58, max_row=58+len(trend_data))
    chart_line.add_data(vals_l, titles_from_data=True)
    chart_line.set_categories(cats_l)
    chart_line.series[0].graphicalProperties.line.width = 28000
    ws.add_chart(chart_line, "B75")

# Column widths
ws.column_dimensions['A'].width = 3
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
    ws.column_dimensions[col].width = 16

# =============================================
# Sheet 2: Raw Data
# =============================================
ws2 = wb.create_sheet("Raw Data")
ws2.sheet_properties.tabColor = "00B894"

headers = ['Rank', 'Title', 'Author', 'Publisher', 'Publish Date', 'Sale Price',
           'Original Price', 'Discount Rate', 'Sale Index', 'Review Count', 'Rating',
           'Detail Link', 'Description', 'Discount Rate Numeric', 'Publish Year', 'Publish Month']

display_headers = ['순위', '제목', '저자', '출판사', '출판일', '판매가',
                   '정가', '할인율', '판매지수', '리뷰 수', '평점',
                   '상세 링크', '소개', '할인율(수치)', '출판년도', '출판월']

for c, header in enumerate(display_headers, 1):
    cell = ws2.cell(row=1, column=c, value=header)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = THIN_BORDER

for r_idx, (_, row_data) in enumerate(df.iterrows(), 2):
    for c_idx, col in enumerate(headers, 1):
        val = row_data[col]
        if pd.isna(val):
            val = ''
        cell = ws2.cell(row=r_idx, column=c_idx, value=val)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        if r_idx % 2 == 0:
            cell.fill = ALT_FILL
        if col == 'Sale Price' or col == 'Original Price':
            cell.number_format = MONEY_FORMAT
        elif col == 'Sale Index':
            cell.number_format = MONEY_FORMAT
        elif col in ('Rank', 'Review Count', 'Publish Year', 'Publish Month'):
            cell.number_format = '0'
        elif col == 'Rating':
            cell.number_format = '0.0'
        elif col == 'Discount Rate Numeric':
            cell.number_format = '0.0'

ws2.column_dimensions['A'].width = 6
ws2.column_dimensions['B'].width = 45
ws2.column_dimensions['C'].width = 20
ws2.column_dimensions['D'].width = 18
ws2.column_dimensions['E'].width = 14
ws2.column_dimensions['F'].width = 12
ws2.column_dimensions['G'].width = 12
ws2.column_dimensions['H'].width = 10
ws2.column_dimensions['I'].width = 12
ws2.column_dimensions['J'].width = 10
ws2.column_dimensions['K'].width = 8
ws2.column_dimensions['L'].width = 30
ws2.column_dimensions['M'].width = 50
ws2.column_dimensions['N'].width = 14
ws2.column_dimensions['O'].width = 10
ws2.column_dimensions['P'].width = 10

ws2.auto_filter.ref = f"A1:P{len(df)+1}"
ws2.freeze_panes = 'A2'

# Add scatter chart to Dashboard (now that Raw Data exists)
chart_scatter = ScatterChart()
chart_scatter.title = "평점 vs 판매지수"
chart_scatter.x_axis.title = "평점"
chart_scatter.y_axis.title = "판매지수"
chart_scatter.width = 22
chart_scatter.height = 14
xvals_ref = Reference(ws2, min_col=11, min_row=2, max_row=len(df)+1)
yvals_ref = Reference(ws2, min_col=9, min_row=2, max_row=len(df)+1)
chart_scatter.add_data(yvals_ref, titles_from_data=False)
chart_scatter.series[0].xvalues = xvals_ref
chart_scatter.legend = None
ws.add_chart(chart_scatter, "B41")

# =============================================
# Sheet 3: Publisher Analysis
# =============================================
ws3 = wb.create_sheet("Publisher Analysis")
ws3.sheet_properties.tabColor = "FDCB6E"

ws3.merge_cells('A1:F1')
ws3['A1'] = '출판사별 상세 분석'
ws3['A1'].font = TITLE_FONT

pub_analysis = df.groupby('Publisher').agg(
    Count=('Title', 'count'),
    Avg_Price=('Sale Price', 'mean'),
    Avg_Rating=('Rating', 'mean'),
    Avg_Review=('Review Count', 'mean'),
    Avg_Sale_Index=('Sale Index', 'mean'),
    Total_Review=('Review Count', 'sum')
).reset_index().sort_values('Count', ascending=False)

pub_headers = ['출판사', '도서 수', '평균 판매가', '평균 평점', '평균 리뷰 수', '총 리뷰 수']
for c, h in enumerate(pub_headers, 1):
    cell = ws3.cell(row=3, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center')
    cell.border = THIN_BORDER

for i, (_, row_data) in enumerate(pub_analysis.iterrows()):
    r = 4 + i
    ws3.cell(row=r, column=1, value=row_data['Publisher']).font = DATA_FONT
    ws3.cell(row=r, column=2, value=int(row_data['Count'])).font = DATA_FONT
    ws3.cell(row=r, column=3, value=round(row_data['Avg_Price'], 0)).font = DATA_FONT
    ws3.cell(row=r, column=4, value=round(row_data['Avg_Rating'], 2)).font = DATA_FONT
    ws3.cell(row=r, column=5, value=round(row_data['Avg_Review'], 1)).font = DATA_FONT
    ws3.cell(row=r, column=6, value=int(row_data['Total_Review'])).font = DATA_FONT
    for c in range(1, 7):
        ws3.cell(row=r, column=c).border = THIN_BORDER
        ws3.cell(row=r, column=c).alignment = Alignment(horizontal='center')
        if i % 2 == 1:
            ws3.cell(row=r, column=c).fill = ALT_FILL
    ws3.cell(row=r, column=3).number_format = '#,##0'
    ws3.cell(row=r, column=6).number_format = '#,##0'

ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 10
ws3.column_dimensions['C'].width = 14
ws3.column_dimensions['D'].width = 10
ws3.column_dimensions['E'].width = 12
ws3.column_dimensions['F'].width = 12

chart_pub_detail = BarChart()
chart_pub_detail.type = "col"
chart_pub_detail.style = 10
chart_pub_detail.title = "출판사별 도서 수"
chart_pub_detail.width = 22
chart_pub_detail.height = 14
last_row_pub = 3 + len(pub_analysis)
cats_pa = Reference(ws3, min_col=1, min_row=4, max_row=last_row_pub)
vals_pa = Reference(ws3, min_col=2, min_row=3, max_row=last_row_pub)
chart_pub_detail.add_data(vals_pa, titles_from_data=True)
chart_pub_detail.set_categories(cats_pa)
chart_pub_detail.y_axis.title = "도서 수"
ws3.add_chart(chart_pub_detail, "H3")

# =============================================
# Sheet 4: Keyword Search (Manual helper)
# =============================================
ws4 = wb.create_sheet("Keyword Search")
ws4.sheet_properties.tabColor = "E17055"

ws4.merge_cells('A1:C1')
ws4['A1'] = '키워드 검색 가이드'
ws4['A1'].font = TITLE_FONT

ws4['A3'] = '검색 방법'
ws4['A3'].font = Font(name='Arial', bold=True, size=11)
ws4['B3'] = 'Raw Data 시트에서 제목(B열) 또는 소개(M열)에 대해 Excel 필터 또는 Ctrl+F로 키워드 검색'
ws4['B3'].font = DATA_FONT

ws4['A5'] = '자주 쓰는 키워드'
ws4['A5'].font = Font(name='Arial', bold=True, size=11)

keywords = ['클로드', '챗GPT', '제미나이', '바이브 코딩', '에듀테크', '프롬프트', 'AI', '파이썬', '엑셀', '교육']
ws4['A6'] = '키워드'
ws4['B6'] = '포함 도서 수'
ws4['A6'].font = HEADER_FONT
ws4['B6'].font = HEADER_FONT
ws4['A6'].fill = HEADER_FILL
ws4['B6'].fill = HEADER_FILL
ws4['A6'].alignment = Alignment(horizontal='center')
ws4['B6'].alignment = Alignment(horizontal='center')

for i, kw in enumerate(keywords):
    r = 7 + i
    count = df[df['Title'].str.contains(kw, case=False, na=False) | df['Description'].str.contains(kw, case=False, na=False)].shape[0]
    ws4[f'A{r}'] = kw
    ws4[f'B{r}'] = count
    ws4[f'A{r}'].font = DATA_FONT
    ws4[f'B{r}'].font = DATA_FONT
    ws4[f'B{r}'].alignment = Alignment(horizontal='center')
    if i % 2 == 1:
        ws4[f'A{r}'].fill = ALT_FILL
        ws4[f'B{r}'].fill = ALT_FILL

ws4.column_dimensions['A'].width = 20
ws4.column_dimensions['B'].width = 55
ws4.column_dimensions['C'].width = 20

chart_kw = BarChart()
chart_kw.type = "col"
chart_kw.style = 10
chart_kw.title = "키워드별 도서 수"
chart_kw.y_axis.title = "도서 수"
chart_kw.width = 22
chart_kw.height = 14
cats_kw = Reference(ws4, min_col=1, min_row=7, max_row=6+len(keywords))
vals_kw = Reference(ws4, min_col=2, min_row=6, max_row=6+len(keywords))
chart_kw.add_data(vals_kw, titles_from_data=True)
chart_kw.set_categories(cats_kw)
ws4.add_chart(chart_kw, "A18")

wb.save(OUTPUT_PATH)
print(f"대시보드 생성 완료: {OUTPUT_PATH}")
